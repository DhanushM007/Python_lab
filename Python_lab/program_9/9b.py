from openpyxl import Workbook
from openpyxl.styles import Font
wb=Workbook()
sheet=wb.active
sheet.title="Language"
wb.create_sheet(title="capital")

state=["karnataka","TamilNadu","Telagana"]
lang=["kannada","Tamil","Telugu"]
capital=["Bengaluru", "Chennai","Hyderabad"]
code=["KA","TN","TS"]

sheet.cell(row=1,column=1).value="State"
sheet.cell(row=1,column=2).value="Language"
sheet.cell(row=1,column=3).value="Code"

ft=Font(bold=True)
for row in sheet["A1:c1"]:
    for cell in row:
        cell.font=ft
        
for i in range(2,5):
    sheet.cell(row=i,column=1).value=state[i-2]
    sheet.cell(row=i,column=2).value=lang[i-2]
    sheet.cell(row=i,column=3).value=code[i-2]
    
wb.save("C:\\Users\\jaswa\\Desktop\\newxl.xlsx")
    
datasrch=input("enter the code for finding the corresponding language:")
for i in range(2,5):
    data=sheet.cell(row=i,column=3).value
    if data==datasrch:
        print("The corresponding language for the code is",sheet.cell(row=i,column=2).value)
        

wb.close()

